import os
import sys

import pandas as pd
from openpyxl import Workbook, load_workbook
#from openpyxl.worksheet.worksheet import Worksheet
from PyQt5 import QtWidgets
from PyQt5.QtCore import QRegExp
from PyQt5.QtGui import QRegExpValidator
from PyQt5.QtWidgets import QMessageBox

from UiEditCustomer import Ui_editCustomerUi

dfhos = pd.read_excel('calibration_info.xlsx','hospitals')




class EC_Ui(QtWidgets.QMainWindow):
    def __init__(self):
        super().__init__()
        self.ui = Ui_editCustomerUi()
        self.ui.setupUi(self)

        self.ui.cbSelectHospital.addItem("")
        self.ui.cbSelectKod.addItem("")

        

        for i in range(0, len(dfhos.index)):
            self.ui.cbSelectHospital.addItem(str(dfhos.loc[i, "hastane"]))

        for i in range(0, len(dfhos.index)):
            self.ui.cbSelectKod.addItem(str(dfhos.loc[i, "kod"]))
        
        
        self.ui.cbSelectHospital.currentTextChanged.connect(self.isno_kod_birlik_doldur)
        self.ui.cbSelectKod.currentTextChanged.connect(self.isno_birlik_hospital_doldur)
        self.ui.edit_record.clicked.connect(self.edit_customer_record)


    def isno_kod_birlik_doldur(self):
        try:
            hospitalNameCheck = self.ui.cbSelectHospital.currentText()

            if hospitalNameCheck == "":
                self.ui.lineSelectIsno.setText("")
                self.ui.lineSelectBirlik.setText("")
                self.ui.cbSelectKod.setCurrentText("")
            else:
                hospital_birlik_fhEdit = dfhos[dfhos['hastane'] == hospitalNameCheck]["birlik"].iloc[0]
                hospital_isno_fhEdit = dfhos[dfhos['hastane'] == hospitalNameCheck]["isno"].iloc[0]
                hospital_kod_fhEdir = dfhos[dfhos['hastane'] == hospitalNameCheck]["kod"].iloc[0]

                self.ui.lineSelectIsno.setText(str(hospital_isno_fhEdit))
                self.ui.lineSelectBirlik.setText(str(hospital_birlik_fhEdit))
                self.ui.cbSelectKod.setCurrentText(str(hospital_kod_fhEdir))
        except IndexError:
            pass

    def isno_birlik_hospital_doldur(self):
        try:
            hospitalCodeCheck = self.ui.cbSelectKod.currentText()

            if hospitalCodeCheck == "":
                self.ui.lineSelectIsno.setText("")
                self.ui.lineSelectBirlik.setText("")
                self.ui.cbSelectHospital.setCurrentText("")
            else:
                hospital_birlik_fcEdit = dfhos[dfhos['kod'] == hospitalCodeCheck]["birlik"].iloc[0]
                hospital_isno_fcEdit = dfhos[dfhos['kod'] == hospitalCodeCheck]["isno"].iloc[0]
                hospital_name_fcEdit = dfhos[dfhos['kod'] == hospitalCodeCheck]["hastane"].iloc[0]

                self.ui.lineSelectIsno.setText(str(hospital_isno_fcEdit))
                self.ui.lineSelectBirlik.setText(str(hospital_birlik_fcEdit))
                self.ui.cbSelectHospital.setCurrentText(str(hospital_name_fcEdit))
        except IndexError:
            pass
    
        self.ui.editCustomerOrder.setText(str(self.ui.lineSelectIsno.text()))
        self.ui.editCustomerCode.setText(str(self.ui.cbSelectKod.currentText()))
        self.ui.editCustomerAddres.setText(str(self.ui.lineSelectBirlik.text()))
        self.ui.editCustomerName.setText(str(self.ui.cbSelectHospital.currentText()))

    def edit_customer_record(self):

        changing1 = str
        changing2 = str
        changing3 = str

        search_value5 = dfhos["kod"].isin([self.ui.editCustomerCode.text()]).any()

        ındexnumber = dfhos[dfhos["hastane"] == self.ui.cbSelectHospital.currentText()].index.values
        seperateIndex = str(ındexnumber)
        foundnumber = int(seperateIndex.find("[")) + 2
        

        if self.ui.cbSelectKod.currentText() == self.ui.editCustomerCode.text():
            changing1 = ""
        elif self.ui.cbSelectKod.currentText() != self.ui.editCustomerCode.text():
            changing1 = self.ui.cbSelectKod.currentText() + " => " + self.ui.editCustomerCode.text()

        if self.ui.lineSelectBirlik.text() == self.ui.editCustomerAddres.text():
            changing2 = ""
        elif self.ui.lineSelectBirlik.text() != self.ui.editCustomerAddres.text():  
            changing2 = self.ui.lineSelectBirlik.text() + " => " + self.ui.editCustomerAddres.text()

        if self.ui.cbSelectHospital.currentText() == self.ui.editCustomerName.text():
            changing3 = ""
        elif self.ui.cbSelectHospital.currentText() != self.ui.editCustomerName.text():  
            changing3 = self.ui.cbSelectHospital.currentText() + " => " + self.ui.editCustomerName.text()
        
        if any(changing1) == False and any(changing2) == False and any(changing3) == False:
            QMessageBox.information(self, 'UYARI', "DÜZENLEME YAPMADINIZ", QMessageBox.Close)

        #Sadece hastane adı değiştirme
        elif any(changing1) == False and any(changing2) == False and  any(changing3) == True:
            msgboxtext = "DEĞİŞİKLİKLER KAYDEDİLSİN Mİ?\n" + changing3

            btnresponse = QMessageBox.question(self, 'ONAY', msgboxtext, QMessageBox.Yes | QMessageBox.No)
            if btnresponse == QMessageBox.Yes:

                newCustomerName = self.ui.editCustomerName.text()
                calibrationInfoExcel = load_workbook("calibration_info.xlsx")
                worksheet = calibrationInfoExcel.active
                worksheet = calibrationInfoExcel["hospitals"]

                worksheet["C" + str(foundnumber)] = newCustomerName

                calibrationInfoExcel.save("calibration_info.xlsx")
                calibrationInfoExcel.close()
                self.statusBar().showMessage("Müşteri Güncellendi.", 3000)

        #Sadece birlik&adres değiştirme
        elif any(changing1) == False and any(changing2) == True and any(changing3) == False:
            msgboxtext = "DEĞİŞİKLİKLER KAYDEDİLSİN Mİ?\n" + changing2

            btnresponse = QMessageBox.question(self, 'ONAY', msgboxtext, QMessageBox.Yes | QMessageBox.No)
            if btnresponse == QMessageBox.Yes:
               
                newCustomerBirlik = self.ui.editCustomerAddres.text()
                calibrationInfoExcel = load_workbook("calibration_info.xlsx")
                worksheet = calibrationInfoExcel.active
                worksheet = calibrationInfoExcel["hospitals"]

                worksheet["B" + str(foundnumber)] = newCustomerBirlik

                calibrationInfoExcel.save("calibration_info.xlsx")
                calibrationInfoExcel.close()
                self.statusBar().showMessage("Müşteri Güncellendi.", 3000)

        #Birlik ve Adres değiştirme
        elif any(changing1) == False and any(changing2) == True and any(changing3) == True:
            msgboxtext = "DEĞİŞİKLİKLER KAYDEDİLSİN Mİ?\n" + changing2 + "\n" + changing3

            btnresponse = QMessageBox.question(self, 'ONAY', msgboxtext, QMessageBox.Yes | QMessageBox.No)
            if btnresponse == QMessageBox.Yes:
                newCustomerName = self.ui.editCustomerName.text()
                newCustomerBirlik = self.ui.editCustomerAddres.text()
                calibrationInfoExcel = load_workbook("calibration_info.xlsx")
                worksheet = calibrationInfoExcel.active
                worksheet = calibrationInfoExcel["hospitals"]

                worksheet["C" + str(foundnumber)] = newCustomerName
                worksheet["B" + str(foundnumber)] = newCustomerBirlik

                calibrationInfoExcel.save("calibration_info.xlsx")
                calibrationInfoExcel.close()
                self.statusBar().showMessage("Müşteri Güncellendi.", 3000)

        #Sadece hastane kodu düzenleme
        elif any(changing1) == True and any(changing2) == False and any(changing3) == False:

            if search_value5 == True and self.ui.editCustomerCode.text() != self.ui.cbSelectKod.currentText():
                msgboxtext = "MÜŞTERİ KODU BENZERSİZ OLMALI. KAYDETMEYE ÇALIŞTIĞINIZ KOD BAŞKA BİR MÜŞTERİYE AİT!"
                QMessageBox.question(self, 'UYARI', msgboxtext, QMessageBox.Close)
            else:
                msgboxtext = "DEĞİŞİKLİKLER KAYDEDİLSİN Mİ?\n" + changing1 

                btnresponse = QMessageBox.question(self, 'ONAY', msgboxtext, QMessageBox.Yes | QMessageBox.No)
                if btnresponse == QMessageBox.Yes:

                    newCustomerCode = self.ui.editCustomerCode.text()
                    calibrationInfoExcel = load_workbook("calibration_info.xlsx")
                    worksheet = calibrationInfoExcel.active
                    worksheet = calibrationInfoExcel["hospitals"]

                    worksheet["D" + str(foundnumber)] = newCustomerCode

                    calibrationInfoExcel.save("calibration_info.xlsx")
                    calibrationInfoExcel.close()
                    self.statusBar().showMessage("Müşteri Güncellendi.", 3000)
        #Kod ve Hatane düzenleme
        elif any(changing1) == True and any(changing2) == False and any(changing3) == True:

            if search_value5 == True and self.ui.editCustomerCode.text() != self.ui.cbSelectKod.currentText():
                msgboxtext = "MÜŞTERİ KODU BENZERSİZ OLMALI. KAYDETMEYE ÇALIŞTIĞINIZ KOD BAŞKA BİR MÜŞTERİYE AİT!"
                QMessageBox.question(self, 'UYARI', msgboxtext, QMessageBox.Close)
            else:
                msgboxtext = "DEĞİŞİKLİKLER KAYDEDİLSİN Mİ?\n" + changing1 + "\n" + changing3

                btnresponse = QMessageBox.question(self, 'ONAY', msgboxtext, QMessageBox.Yes | QMessageBox.No)
                if btnresponse == QMessageBox.Yes:

                    newCustomerCode = self.ui.editCustomerCode.text()
                    newCustomerName = self.ui.editCustomerName.text()

                    calibrationInfoExcel = load_workbook("calibration_info.xlsx")
                    worksheet = calibrationInfoExcel.active
                    worksheet = calibrationInfoExcel["hospitals"]

                    worksheet["D" + str(foundnumber)] = newCustomerCode
                    worksheet["C" + str(foundnumber)] = newCustomerName

                    calibrationInfoExcel.save("calibration_info.xlsx")
                    calibrationInfoExcel.close()
                    self.statusBar().showMessage("Müşteri Güncellendi.", 3000)

                     
        #Kod ve birlik düzenleme
        elif any(changing1) == True and any(changing2) == True and any(changing3) == False:

            if search_value5 == True and self.ui.editCustomerCode.text() != self.ui.cbSelectKod.currentText():
                msgboxtext = "MÜŞTERİ KODU BENZERSİZ OLMALI. KAYDETMEYE ÇALIŞTIĞINIZ KOD BAŞKA BİR MÜŞTERİYE AİT!"
                QMessageBox.question(self, 'UYARI', msgboxtext, QMessageBox.Close)
            else:
                msgboxtext = "DEĞİŞİKLİKLER KAYDEDİLSİN Mİ?\n" + changing1 + "\n" + changing2
                btnresponse = QMessageBox.question(self, 'ONAY', msgboxtext, QMessageBox.Yes | QMessageBox.No)
                if btnresponse == QMessageBox.Yes:

                    newCustomerCode = self.ui.editCustomerCode.text()
                    newCustomerBirlik = self.ui.editCustomerAddres.text()

                    calibrationInfoExcel = load_workbook("calibration_info.xlsx")
                    worksheet = calibrationInfoExcel.active
                    worksheet = calibrationInfoExcel["hospitals"]

                    worksheet["D" + str(foundnumber)] = newCustomerCode
                    worksheet["B" + str(foundnumber)] = newCustomerBirlik

                    calibrationInfoExcel.save("calibration_info.xlsx")
                    calibrationInfoExcel.close()
                    self.statusBar().showMessage("Müşteri Güncellendi.", 3000)

        #HEPSİNİ DÜZENLEME
        elif any(changing1) == True and any(changing2) == True and any(changing3) == True:

            if search_value5 == True and self.ui.editCustomerCode.text() != self.ui.cbSelectKod.currentText():
                msgboxtext = "MÜŞTERİ KODU BENZERSİZ OLMALI. KAYDETMEYE ÇALIŞTIĞINIZ KOD BAŞKA BİR MÜŞTERİYE AİT!"
                QMessageBox.question(self, 'UYARI', msgboxtext, QMessageBox.Close)
            else:
                msgboxtext = "DEĞİŞİKLİKLER KAYDEDİLSİN Mİ?\n" + changing1 + "\n" + changing2 + "\n" + changing3
                btnresponse = QMessageBox.question(self, 'ONAY', msgboxtext, QMessageBox.Yes | QMessageBox.No)
                if btnresponse == QMessageBox.Yes:

                    newCustomerCode = self.ui.editCustomerCode.text()
                    newCustomerBirlik = self.ui.editCustomerAddres.text()
                    newCustomerName = self.ui.editCustomerName.text()

                    calibrationInfoExcel = load_workbook("calibration_info.xlsx")
                    worksheet = calibrationInfoExcel.active
                    worksheet = calibrationInfoExcel["hospitals"]

                    worksheet["D" + str(foundnumber)] = newCustomerCode
                    worksheet["B" + str(foundnumber)] = newCustomerBirlik
                    worksheet["C" + str(foundnumber)] = newCustomerName

                    calibrationInfoExcel.save("calibration_info.xlsx")
                    calibrationInfoExcel.close()
                    self.statusBar().showMessage("Müşteri Güncellendi.", 3000)

                  


        
