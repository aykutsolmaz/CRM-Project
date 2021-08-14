"""
This code written by AYKUT SOLMAZ to record calibration information, create certificate and record data history.

AYKUT SOLMAZ
Biomedical Engineer

Project CRP (Calibration Record Managing)

"""

from typing import Text
from PyQt5 import QtCore
import pandas as pd
import sys
import os
import openpyxl
import csv
import time
import locale

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from PyQt5 import QtWidgets, QtTest
from PyQt5.QtGui import *
from PyQt5.QtCore import QDateTime, QT_TR_NOOP_UTF8, Qt, QDate, QTime, QPropertyAnimation, QTimeLine, QRect, QRegExp
from csv import writer, reader
from PyQt5.QtWidgets import QMessageBox, QStyleFactory

from mainPage import Ui_mainWindow

from addNewCustomer import ANC_ui
from editCustomer import EC_Ui
from editStandartCertificateWay import ESCW_Ui
from editRecordFileWay import ERFW_Ui

dfhos = pd.read_excel('calibration_info.xlsx','hospitals')
dfhoscode = pd.read_excel('calibration_info.xlsx','hospitals')
dfdevice = pd.read_excel('calibration_info.xlsx', 'device')
dfdep = pd.read_excel('calibration_info.xlsx', 'department')
dfbrand = pd.read_excel('calibration_info.xlsx', 'brand')
dfmodel = pd.read_excel('calibration_info.xlsx', 'model')
dfcalibperson = pd.read_excel('calibration_info.xlsx', 'calibperson')

thisSetupPath = os.getcwd()

locale.setlocale(locale.LC_NUMERIC, 'en_DK.UTF-8')
  
class MP_Ui(QtWidgets.QMainWindow):
    def __init__(self):
        super(MP_Ui, self).__init__()
        self.ui = Ui_mainWindow()
        self.ui.setupUi(self)

        self.addnewcus = ANC_ui() 
        self.editcurrentcustomer = EC_Ui()
        self.editStandartCertificateWay = ESCW_Ui()
        self.editRecordFileWay = ERFW_Ui()
        
        self.ui.CommentBox.setReadOnly(True)
        self.ui.CommentBoxSerial.setReadOnly(True)
    
        self.ui.pageOnebyOne = 0
        self.ui.pageSerial = 1

        self.ui.emptyPage = 0
        self.ui.ameliyatlambasi = 1
        self.ui.sicaklikolcum = 2
        self.ui.defibrilator = 3

        # Combobox a ilk boş seçenek vermek için
        self.ui.cb_hospital.addItem("")
        self.ui.cb_hospitalcode.addItem("")
        self.ui.cb_device.addItem("")
        self.ui.cb_department.addItem("")
        self.ui.cb_brand.addItem("")
        self.ui.cb_model.addItem("")
        self.ui.cb_calibperson.addItem("")
        self.ui.cb_calibpersonSerial.addItem("")
         
        
        #self.ui.kn_no.setValidator(QIntValidator(0, 2147483647,self))

        sıcaklıkValidator = QRegExpValidator(QRegExp("[-+]?[0,0-99,9]+[-,]"))

        self.ui.celciusValue.setValidator(sıcaklıkValidator)
        
        
        self.ui.tag_code_end.setValidator(QIntValidator(0, 2147483647,self))
        
        #hastane bigilerini combobox a ekleme
        for i in range(0, len(dfhos.index)):
            self.ui.cb_hospital.addItem(str(dfhos.loc[i, "hastane"]))
        
        #hastane kodu bigilerini combobox a ekleme
        for i in range(0, len(dfhos.index)):
            self.ui.cb_hospitalcode.addItem(str(dfhoscode.loc[i, "kod"]))
            
        #cihaz bigilerini combobox a ekleme
        for i in range(0, len(dfdevice.index)):
            self.ui.cb_device.addItem(str(dfdevice.loc[i, "device"]))
            
        #bölüm bigilerini combobox a ekleme
        for i in range(0, len(dfdep.index)):
            self.ui.cb_department.addItem(str(dfdep.loc[i,"department"]))
        
        #marka bigilerini combobox a ekleme
        for i in range(0, len(dfbrand.index)):
            self.ui.cb_brand.addItem(str(dfbrand.loc[i,"brand"]))
            
        #model bilgilerini combobox a ekleme
        for i in range(0, len(dfmodel.index)):
            self.ui.cb_model.addItem(str(dfmodel.loc[i, "model"]))

        #kalibrasyonu yapan bigilerini combobox a ekleme
        for i in range(0, len(dfcalibperson.index)):
            self.ui.cb_calibperson.addItem(str(dfcalibperson.loc[i,"calibperson"]))
            self.ui.cb_calibpersonSerial.addItem(str(dfcalibperson.loc[i,"calibperson"]))

        self.ui.addCustomer.triggered.connect(self.openAddNewCus) 
        self.ui.editCustomer.triggered.connect(self.openEditCus)
        self.ui.editSCW.triggered.connect(self.openEditStandartCertificateWay)
        self.ui.editRFW.triggered.connect(self.openEditRecordFileWay)

        self.ui.cb_hospital.currentTextChanged.connect(self.birlik_isno_kod_doldur)
        self.ui.cb_hospitalcode.currentTextChanged.connect(self.birlik_isno_hospital_doldur)
        self.ui.dateEdit.dateChanged.connect(self.birlik_isno_kod_doldur)
        self.ui.add_device.clicked.connect(self.add_device)
        self.ui.add_department.clicked.connect(self.add_dep)
        self.ui.add_brand.clicked.connect(self.add_brand)
        self.ui.add_model.clicked.connect(self.add_model)

        self.ui.add_record.clicked.connect(self.add_record_to_excel)
        self.ui.add_recordSerial.clicked.connect(self.add_record_to_excel_Serial)

        self.ui.tag_code_end.textChanged.connect(self.certificate_code_edit)
        self.ui.tag_code_endSerial.textChanged.connect(self.certificate_code_edit)
        self.ui.cb_hospital.currentTextChanged.connect(self.certificate_code_edit)
        self.ui.cb_hospitalcode.currentTextChanged.connect(self.certificate_code_edit) 

        self.ui.checkBoxComment.stateChanged.connect(self.activateCommentBox)
        self.ui.checkBoxCommentSerial.stateChanged.connect(self.activateCommentBox)
        self.ui.checkBoxHospitalFix.stateChanged.connect(self.deactivateHospitalChoice)
        self.ui.checkBoxSelectSerial.stateChanged.connect(self.selectProductionPage)

        self.ui.cb_device.currentTextChanged.connect(self.getValueEnterPage)

        self.ui.cb_device.currentTextChanged.connect(self.slideaddRecordFrame)
        self.ui.cb_device.currentTextChanged.connect(self.setWidthEnterValueFrame)

        self.ui.listWidget.doubleClicked.connect(self.showDetailofData)
        
        now = QDate.currentDate()  
        self.ui.dateEdit.setDate(now)
    
    
    def birlik_isno_hospital_doldur(self):

        xyz = self.ui.cb_hospitalcode.currentText()

        try:
            if xyz == "":
                self.ui.line_birlik.setText("")
                self.ui.line_isno.setText("")
                self.ui.cb_hospital.setCurrentText("")
                self.ui.tag_code_front.setText("")
                self.ui.tag_code_end.setText("")

                self.ui.cb_device.setEnabled(False)
                self.ui.cb_department.setEnabled(False)
                self.ui.cb_brand.setEnabled(False)
                self.ui.cb_model.setEnabled(False)
                self.ui.cb_device.setEnabled(False)
                self.ui.seri_no.setEnabled(False)
                self.ui.kn_no.setEnabled(False)
                self.ui.add_device.setEnabled(False)
                self.ui.add_department.setEnabled(False)
                self.ui.add_brand.setEnabled(False)
                self.ui.add_model.setEnabled(False)
                self.ui.tag_code_end.setEnabled(False)
                self.ui.cb_calibperson.setEnabled(False) 
                self.ui.add_record.setEnabled(False)

                self.ui.tag_code_endSerial.setEnabled(False)
                self.ui.tag_code_to_endSerial.setEnabled(False)
                self.ui.cb_calibpersonSerial.setEnabled(False)
                self.ui.add_recordSerial.setEnabled(False)
            else:
                hospital_birlik_fc = dfhoscode[dfhoscode['kod'] == xyz]["birlik"].iloc[0]
                hospital_isno_fc = dfhoscode[dfhoscode['kod'] == xyz]["isno"].iloc[0]
                hospital_name_fc = dfhoscode[dfhoscode['kod'] == xyz]["hastane"].iloc[0]

                self.ui.line_birlik.setText(str(hospital_birlik_fc))
                self.ui.cb_hospital.setCurrentText(str(hospital_name_fc))

                if  len(str(hospital_isno_fc)) == 2:
                    self.ui.line_isno.setText("0" + str(hospital_isno_fc))
                else:
                    self.ui.line_isno.setText(str(hospital_isno_fc))                

                date_text = self.ui.dateEdit.text()
                month = date_text[3:5]
                year = date_text[8:10]

                self.ui.tag_code_front.setText(month + year + self.ui.cb_hospitalcode.currentText())
                self.ui.tag_code_frontSerial.setText(month + year + self.ui.cb_hospitalcode.currentText())

                chose_order_number = month + year + self.ui.cb_hospitalcode.currentText() + self.ui.line_isno.text()

                try:
                    self.ui.listWidget.clear()
                    with open(thisSetupPath + "\\history\\" + chose_order_number + ".csv", encoding='utf-8') as file:
                        csv_reader = csv.DictReader(file)
                        for row  in csv_reader:
                            tagCode = f'{row["Etiket No"]}'
                            self.ui.listWidget.addItems([tagCode])
                            self.ui.listWidget.scrollToBottom()
                except FileNotFoundError:
                    with open(thisSetupPath + "\\history\\" + chose_order_number + ".csv", "w", newline= "", encoding='utf-8') as file:
                        csv_writer = csv.writer(file)
                        csv_writer.writerow(["Zaman","Bulundugu Bolum","Cihaz Adi","Etiket No","Seri No","Kunye No"])
                        print(os.getcwd())

                            
                self.ui.cb_device.setEnabled(True)
                self.ui.cb_department.setEnabled(True)
                self.ui.cb_brand.setEnabled(True)
                self.ui.cb_model.setEnabled(True)
                self.ui.cb_device.setEnabled(True)
                self.ui.seri_no.setEnabled(True)
                self.ui.kn_no.setEnabled(True)
                self.ui.add_device.setEnabled(True)
                self.ui.add_department.setEnabled(True)
                self.ui.add_brand.setEnabled(True)
                self.ui.add_model.setEnabled(True)
                self.ui.tag_code_end.setEnabled(True)
                self.ui.cb_calibperson.setEnabled(True)
                self.ui.add_record.setEnabled(True)

                self.ui.tag_code_endSerial.setEnabled(True)
                self.ui.tag_code_to_endSerial.setEnabled(True)
                self.ui.cb_calibpersonSerial.setEnabled(True)
                self.ui.add_recordSerial.setEnabled(True)

        except IndexError:
            pass

    def birlik_isno_kod_doldur(self):
        abc = self.ui.cb_hospital.currentText() 
        try:
            if abc == "":
                self.ui.line_birlik.setText("")
                self.ui.line_isno.setText("")
                self.ui.cb_hospitalcode.setCurrentText("")
                self.ui.tag_code_front.setText("")
                self.ui.tag_code_end.setText("")

                self.ui.cb_device.setEnabled(False)
                self.ui.cb_department.setEnabled(False)
                self.ui.cb_brand.setEnabled(False)
                self.ui.cb_model.setEnabled(False)
                self.ui.cb_device.setEnabled(False)
                self.ui.seri_no.setEnabled(False)
                self.ui.kn_no.setEnabled(False)
                self.ui.add_device.setEnabled(False)
                self.ui.add_department.setEnabled(False)
                self.ui.add_brand.setEnabled(False)
                self.ui.add_model.setEnabled(False)
                self.ui.tag_code_end.setEnabled(False)
                self.ui.cb_calibperson.setEnabled(False)
                self.ui.add_record.setEnabled(False) 

                self.ui.tag_code_endSerial.setEnabled(False)
                self.ui.tag_code_to_endSerial.setEnabled(False)
                self.ui.cb_calibpersonSerial.setEnabled(False)
                self.ui.add_recordSerial.setEnabled(False)
            else:                
                hospital_birlik_fh = dfhos[dfhos['hastane'] == abc]["birlik"].iloc[0]
                hospital_isno_fh = dfhos[dfhos['hastane'] == abc]["isno"].iloc[0]
                hospital_kod_fh = dfhos[dfhos['hastane'] == abc]["kod"].iloc[0]

                self.ui.line_birlik.setText(str(hospital_birlik_fh))
                self.ui.cb_hospitalcode.setCurrentText(str(hospital_kod_fh)) 

                if  len(str(hospital_isno_fh)) == 2:
                    self.ui.line_isno.setText("0" + str(hospital_isno_fh))
                else:
                    self.ui.line_isno.setText(str(hospital_isno_fh))

                date_text = self.ui.dateEdit.text()
                month = date_text[3:5]
                year = date_text[8:10]

                self.ui.tag_code_front.setText(month + year + self.ui.cb_hospitalcode.currentText())
                self.ui.tag_code_frontSerial.setText(month + year + self.ui.cb_hospitalcode.currentText())

                chose_order_number = month + year + self.ui.cb_hospitalcode.currentText() + self.ui.line_isno.text()

                try:
                    self.ui.listWidget.clear()
                    with open(thisSetupPath + "\\history\\" + chose_order_number + ".csv", encoding='utf-8') as file:
                        csv_reader = csv.DictReader(file)

                        for row  in csv_reader:
                            tagCode = f'{row["Etiket No"]}'
                            self.ui.listWidget.addItems([tagCode])
                            self.ui.listWidget.scrollToBottom()
                except FileNotFoundError:
                    with open(thisSetupPath + "\\history\\" + chose_order_number + ".csv", "w", newline= "", encoding='utf-8') as file:
                        csv_writer = csv.writer(file)
                        csv_writer.writerow(["Zaman","Bulundugu Bolum","Cihaz Adi","Etiket No","Seri No","Kunye No"])
 
            
                self.ui.cb_device.setEnabled(True)
                self.ui.cb_department.setEnabled(True)
                self.ui.cb_brand.setEnabled(True)
                self.ui.cb_model.setEnabled(True)
                self.ui.cb_device.setEnabled(True)
                self.ui.seri_no.setEnabled(True)
                self.ui.kn_no.setEnabled(True)
                self.ui.add_device.setEnabled(True)
                self.ui.add_department.setEnabled(True)
                self.ui.add_brand.setEnabled(True)
                self.ui.add_model.setEnabled(True)
                self.ui.tag_code_end.setEnabled(True) 
                self.ui.cb_calibperson.setEnabled(True)
                self.ui.add_record.setEnabled(True)

                self.ui.tag_code_endSerial.setEnabled(True)
                self.ui.tag_code_to_endSerial.setEnabled(True)
                self.ui.cb_calibpersonSerial.setEnabled(True)
                self.ui.add_recordSerial.setEnabled(True)  

        except IndexError:
            pass

    def certificate_code_edit(self):

        date_text = self.ui.dateEdit.text()
        month = date_text[3:5]
        year = date_text[8:10]
        textinfo = self.ui.tag_code_end.text()
        textinfoSerial_Begin = self.ui.tag_code_endSerial.text()


        if len(textinfo) == 0: 
            self.ui.tag_code_front.setText(month + year + self.ui.cb_hospitalcode.currentText())
        elif len(textinfo) == 1:
            self.ui.tag_code_front.setText(month + year + self.ui.cb_hospitalcode.currentText() + "000")
        elif len(textinfo) == 2:
            self.ui.tag_code_front.setText(month + year + self.ui.cb_hospitalcode.currentText() + "00")
        elif len(textinfo) == 3:
            self.ui.tag_code_front.setText(month + year + self.ui.cb_hospitalcode.currentText() + "0")
        elif len(textinfo) == 4:
            self.ui.tag_code_front.setText(month + year + self.ui.cb_hospitalcode.currentText() + "")
    
        if len(textinfoSerial_Begin) == 0:
            self.ui.tag_code_frontSerial.setText(month + year + self.ui.cb_hospitalcode.currentText())
        elif len(textinfoSerial_Begin) == 1:
            self.ui.tag_code_frontSerial.setText(month + year + self.ui.cb_hospitalcode.currentText() + "000")
        elif len(textinfoSerial_Begin) == 2:
            self.ui.tag_code_frontSerial.setText(month + year + self.ui.cb_hospitalcode.currentText() + "00")
        elif len(textinfoSerial_Begin) == 3:
            self.ui.tag_code_frontSerial.setText(month + year + self.ui.cb_hospitalcode.currentText() + "0")
        elif len(textinfoSerial_Begin) == 4:
            self.ui.tag_code_frontSerial.setText(month + year + self.ui.cb_hospitalcode.currentText() + "")

    def add_device(self):
        search_value0 = dfdevice.isin([self.ui.cb_device.currentText()]).any().iloc[0]
        
        if search_value0 == True:
            self.ui.statusbar.showMessage("Cihaz Adı: Eklemeye çalışılan cihaz adı zaten var...",2000)
        elif self.ui.cb_device.currentText() == "":
            self.ui.statusbar.showMessage("Cihaz Adı: Boş kayıt ekleyemezsiniz...",2000)
        else:

            dfadd_device = pd.DataFrame({'device': [self.ui.cb_device.currentText()]})
            workbook = load_workbook("calibration_info.xlsx")
            reader = pd.read_excel(r'calibration_info.xlsx',sheet_name='device')

            writer = pd.ExcelWriter('calibration_info.xlsx', engine='openpyxl') 
            writer.book = workbook
            writer.sheets = dict((ws.title, ws) for ws in workbook.worksheets)
            
            dfadd_device.to_excel(writer,index=False,header=False,startrow=len(reader)+1,sheet_name='device')
            writer.close()
            self.ui.statusbar.showMessage("Cihaz Adı: Yeni cihaz eklendi...",2000)

    def add_dep(self):
        search_value1 = dfdep.isin([self.ui.cb_department.currentText()]).any().iloc[0]
        
        if search_value1 == True:
            self.ui.statusbar.showMessage("Bulunduğu Yer: Eklemeye çalışılan bölüm adı zaten var...",2000)
        elif self.ui.cb_department.currentText() == "":
            self.ui.statusbar.showMessage("Bulunduğu Yer: Boş kayıt ekleyemezsiniz...",2000)
        else:

            dfadd_dep = pd.DataFrame({'department': [self.ui.cb_department.currentText()]})
            workbook = load_workbook("calibration_info.xlsx")
            reader = pd.read_excel(r'calibration_info.xlsx',sheet_name='department')

            writer = pd.ExcelWriter('calibration_info.xlsx', engine='openpyxl')
            writer.book = workbook
            writer.sheets = dict((ws.title, ws) for ws in workbook.worksheets)
            
            dfadd_dep.to_excel(writer,index=False,header=False,startrow=len(reader)+1,sheet_name='department')
            writer.close()
            self.ui.statusbar.showMessage("Bulunduğu Yer: Yeni bölüm eklendi...",2000)  

    def add_brand(self):
        search_value2 = dfbrand.isin([self.ui.cb_brand.currentText()]).any().iloc[0]
        
        if search_value2 == True:
            self.ui.statusbar.showMessage("Marka: Eklemeye çalışılan marka adı zaten var...",2000)
        elif self.ui.cb_brand.currentText() == "":
            self.ui.statusbar.showMessage("Marka: Boş kayıt ekleyemezsiniz...",2000)   
        else:

            dfadd_brand = pd.DataFrame({'brand': [self.ui.cb_brand.currentText()]})
            workbook = load_workbook("calibration_info.xlsx")
            reader = pd.read_excel(r'calibration_info.xlsx',sheet_name='brand')

            writer = pd.ExcelWriter('calibration_info.xlsx', engine='openpyxl')
            writer.book = workbook
            writer.sheets = dict((ws.title, ws) for ws in workbook.worksheets)
            
            dfadd_brand.to_excel(writer,index=False,header=False,startrow=len(reader)+1,sheet_name='brand')
            writer.close()
            self.ui.statusbar.showMessage("Marka: Yeni marka eklendi...",2000)

    def add_model(self):
        search_value3 = dfmodel.isin([str(self.ui.cb_model.currentText())]).any().iloc[0]
        
        if search_value3 == True:
            self.ui.statusbar.showMessage("Model: Eklemeye çalışılan model adı zaten var...",2000)
        elif self.ui.cb_model.currentText() == "":
            self.ui.statusbar.showMessage("Model: Boş kayıt ekleyemezsiniz...",2000)  
        else:

            dfadd_model = pd.DataFrame({'model': [self.ui.cb_model.currentText()]})
            workbook = load_workbook("calibration_info.xlsx")
            reader = pd.read_excel(r'calibration_info.xlsx',sheet_name='model')

            writer = pd.ExcelWriter('calibration_info.xlsx', engine='openpyxl')
            writer.book = workbook
            writer.sheets = dict((ws.title, ws) for ws in workbook.worksheets)
            
            dfadd_model.to_excel(writer,index=False,header=False,startrow=len(reader)+1,sheet_name='model')
            writer.close()
            self.ui.statusbar.showMessage("Model: Yeni model eklendi...",2000)

    def getValueEnterPage(self):
        deviceSelection = self.ui.cb_device.currentText()

        if deviceSelection == "AMELİYAT LAMBASI ( LED )":
            self.ui.deviceStackedWidget.setCurrentIndex(1)
        elif deviceSelection == "AMELİYAT LAMBASI":
            self.ui.deviceStackedWidget.setCurrentIndex(1)
        elif deviceSelection == "AŞI TAŞIMA ÇANTASI":
            self.ui.deviceStackedWidget.setCurrentIndex(2)
        elif deviceSelection == "BEDEN DERECESİ":
            self.ui.deviceStackedWidget.setCurrentIndex(2)
        elif deviceSelection == "BENMARİ":
            self.ui.deviceStackedWidget.setCurrentIndex(2)
        elif deviceSelection == "BUZDOLABI":
            self.ui.deviceStackedWidget.setCurrentIndex(2)
        elif deviceSelection == "COLDPACK":
            self.ui.deviceStackedWidget.setCurrentIndex(2)
        elif deviceSelection == "DATA LOGGER":
            self.ui.deviceStackedWidget.setCurrentIndex(2)
        elif deviceSelection == "DATA LOGGER_":
            self.ui.deviceStackedWidget.setCurrentIndex(2)
        elif deviceSelection == "DERİN DONDURUCU":
            self.ui.deviceStackedWidget.setCurrentIndex(2)
        elif deviceSelection == "DEFİBRİLATÖR":
            self.ui.deviceStackedWidget.setCurrentIndex(3)
        elif deviceSelection == "DEFİBRİLATÖR_":
            self.ui.deviceStackedWidget.setCurrentIndex(3)
        else:
            self.ui.deviceStackedWidget.setCurrentIndex(0)
    
    def activateCommentBox(self):
        if self.ui.checkBoxComment.isChecked() == False:
            self.ui.CommentBox.setReadOnly(True)
        elif self.ui.checkBoxComment.isChecked() == True:
            self.ui.CommentBox.setReadOnly(False)

        if self.ui.checkBoxCommentSerial.isChecked() == False:
            self.ui.CommentBoxSerial.setReadOnly(True)
        elif self.ui.checkBoxCommentSerial.isChecked() == True:
            self.ui.CommentBoxSerial.setReadOnly(False)

    def deactivateHospitalChoice(self):
        if self.ui.checkBoxHospitalFix.isChecked() == True:
            self.ui.cb_hospitalcode.setEnabled(False)
            self.ui.cb_hospital.setEnabled(False)
            self.ui.line_isno.setEnabled(False)
            self.ui.line_birlik.setEnabled(False)
        elif self.ui.checkBoxHospitalFix.isChecked() == False:
            self.ui.cb_hospitalcode.setEnabled(True)
            self.ui.cb_hospital.setEnabled(True)
            self.ui.line_isno.setEnabled(True)
            self.ui.line_birlik.setEnabled(True)

    def selectProductionPage(self):
        if self.ui.checkBoxSelectSerial.isChecked() == False:
            self.ui.stackedWidget.setCurrentIndex(0)

        elif self.ui.checkBoxSelectSerial.isChecked() == True:
            self.ui.stackedWidget.setCurrentIndex(1)
   
    def add_record_to_excel(self):

        date_text = self.ui.dateEdit.text()
        month = date_text[3:5]
        year = date_text[8:10]

        time = QDateTime.currentDateTime()
        chosen_hospital = self.ui.cb_hospital.currentText()
        chosen_birlik = self.ui.line_birlik.text()
        chosen_device = self.ui.cb_device.currentText()
        chosen_department = self.ui.cb_department.currentText()
        chosen_brand = self.ui.cb_brand.currentText()
        chosen_model = self.ui.cb_model.currentText()
        chosen_serialnum = self.ui.seri_no.text()
        chosen_knyeno = self.ui.kn_no.text()
        chosen_date = self.ui.dateEdit.text()
        chosen_calpeson = self.ui.cb_calibperson.currentText()
        chosen_certificate_number = self.ui.tag_code_front.text() + self.ui.tag_code_end.text()
        chose_order_number = month + year + self.ui.cb_hospitalcode.currentText() + self.ui.line_isno.text()


        az1Lux = self.ui.lxValue.text()
        az1Sıcaklık = self.ui.celciusValue.text()
        az1MaxPower = self.ui.maxPower.text()

        az2ChargeTime = self.ui.chargeTime.text()

        az3DelayTime = self.ui.delayTime.text()

                   
        def devamke():
            if self.ui.checkBoxComment.isChecked() == True and self.ui.CommentBox.toPlainText() == "":
                optionalCommand = ""
            elif self.ui.checkBoxComment.isChecked() == False:
                optionalCommand = ""
            else:
                optionalCommand = "-" + self.ui.CommentBox.toPlainText()

            if chosen_hospital == "" or chosen_birlik == "" or chosen_device == "" or chosen_department == "" or chosen_brand == "" or chosen_model == "" or chosen_serialnum == "" or chosen_knyeno == "" or chosen_date == "" or chosen_calpeson == "":
                self.statusBar().showMessage("Bilgiler eksik, Gerekli bilgileri giriniz...", 2000) 
            elif self.ui.tag_code_end.text() == "":
                self.statusBar().showMessage("Serifika numarasını giriniz...", 2000)
            else:
                with open('pathFolder.txt', 'r', encoding="utf-8") as file:
                    pathSCW = file.read().splitlines()
                
                try:    
                    standart_certificate = load_workbook(pathSCW[0]+ "\\" + chosen_device + ".xlsx")
                    worksheet = standart_certificate.active
                    worksheet = standart_certificate["Veri"]
                    worksheet ['O1'] = chosen_hospital
                    worksheet ['O2'] = chosen_birlik
                    worksheet ['O2'] = chose_order_number
                    worksheet ['O10'] = chosen_department
                    worksheet ['O5'] = chosen_brand
                    worksheet ['O6'] = chosen_model
                    worksheet ['O7'] = chosen_serialnum
                    worksheet ['O11'] = chosen_knyeno
                    worksheet ['O5'] = chosen_brand
                    worksheet ['O8'] = chosen_date
                    worksheet ['O15'] = chosen_calpeson
                    worksheet ['O12'] = chosen_certificate_number
                    
                    try:
                        az1List = [az1Lux, az1Sıcaklık, az1MaxPower]
                        az1Value = list(filter(lambda x: x != "", az1List))[0]
                    except IndexError:
                        az1Value = ""
                    
                    try:
                        az2list = [az2ChargeTime]
                        az2Value = list(filter(lambda x: x != "", az2list))[0]
                    except IndexError:
                        az2Value = ""

                    try:
                        az3list = [az3DelayTime]
                        az3Value = list(filter(lambda x: x != "", az3list))[0]
                    except IndexError:
                        az3Value = ""
                    
                    try:
                        worksheet ['AZ1'] = locale.atof(az1Value)
                    except ValueError:
                        pass
                    
                    try:
                        worksheet ['AZ2'] = locale.atof(az2Value)
                    except ValueError:
                        pass

                    try:
                        worksheet ['AZ3'] = locale.atof(az3Value)
                    except ValueError:
                        pass



                    try:
                        standart_certificate.save(pathSCW[1]+ "\\" + chosen_department + "-" + chosen_device + "-" + chosen_certificate_number +  optionalCommand +  ".xlsx")
                        standart_certificate.close()
                        self.ui.statusbar.showMessage(chosen_certificate_number + ": Sertifika numaralı kayıt eklendi.",2000)

                        with open (thisSetupPath + "\\history\\" + chose_order_number +".csv", "a", newline='', encoding='utf-8') as file:
                            csv_writer = writer(file)                  
                            csv_writer.writerow([time.toString(), chosen_department, chosen_device, chosen_certificate_number, chosen_serialnum, chosen_knyeno])
                            self.ui.listWidget.scrollToBottom()

                        self.ui.add_record.setEnabled(False)
                        QtTest.QTest.qWait(2000)
                        self.ui.listWidget.clear()

                        with open(thisSetupPath + "\\history\\" + chose_order_number + ".csv", encoding='utf-8') as file:
                            csv_reader = csv.DictReader(file)

                            for row  in csv_reader:
                                tagCode = f'{row["Etiket No"]}'
                                
                                self.ui.listWidget.addItems([tagCode])
                                self.ui.listWidget.scrollToBottom()

                        self.ui.add_record.setEnabled(True)
                        currentTagNumber = int(self.ui.tag_code_end.text())
                        nextTagNumber = currentTagNumber + 1
                        self.ui.tag_code_end.setText(str(nextTagNumber))
                        self.ui.celciusValue.clear()
                        self.ui.lxValue.clear()
                        self.ui.maxPower.clear()
                        self.ui.chargeTime.clear()
                        self.ui.delayTime.clear()
                        if self.ui.checkBoxCommentFix.isChecked() == False:
                            self.ui.CommentBox.clear()
                        else:
                            pass
                    except FileNotFoundError:
                        msgboxtext = "KAYIT konumu bulunamıyor... Dosya yolunu kontrol ediniz."
                        QMessageBox.warning(self, 'UYARI', msgboxtext, QMessageBox.Close)
  
                except FileNotFoundError:
                    msgboxtext = "Standart Sertifika bulunamıyor.\nDosya yolunu kontrol ediniz veya ilgili cihaza ait Standart Sertifika ekleyiniz."
                    QMessageBox.warning(self, 'UYARI', msgboxtext, QMessageBox.Close)

        
        with open('pathFolder.txt', 'r', encoding="utf-8") as file:
            pathSCW = file.read().splitlines()

        df = pd.read_csv(thisSetupPath + "\\history\\" + chose_order_number +".csv",encoding = 'utf-8', engine ='python')
        dfTagNumner = df["Etiket No"]
        dfDepartment = df["Bulundugu Bolum"]
        dfDeviceName = df["Cihaz Adi"]
        dfKunyeNo = df["Kunye No"]
        
        searchTagNumber = dfTagNumner.isin([str(chosen_certificate_number)]).any()
        searchDepartment = dfDepartment.isin([str(chosen_department)]).any()
        searchDeviceName = dfDeviceName.isin([str(chosen_device)]).any()
        try:
            if self.ui.kn_no.text() == "-":
                searchKunyeNo = False
            else:
                searchKunyeNo = dfKunyeNo.isin([int(chosen_knyeno)]).any()
        except ValueError:
                searchKunyeNo = False
                
        if searchTagNumber == False and searchKunyeNo == True: #Sadece Künye No aynıysa
            with open (thisSetupPath + "\\history\\" + chose_order_number +".csv", "r", newline='', encoding='utf-8') as file:
                csv_reader = csv.reader(file)
                matchedCertificateNo = []
                for row in csv_reader:
                    if chosen_knyeno == row[5]:
                        matchedCertificateNo.append(row[3])
            
            matchedCertificateNo = [str(x) for x in matchedCertificateNo]
            msgMatch = ', '.join(matchedCertificateNo)

            msgboxtext = "'" + chosen_knyeno + "'" + " künye numarası kullanılıyor.\nKaydetmek isteyor musunuz?\nEşleşen sertifika numarası: " + msgMatch
            btnresponse = QMessageBox.question(self, "ONAY", msgboxtext, QMessageBox.Yes | QMessageBox.No )
            if btnresponse == QMessageBox.Yes:
                devamke()
               
        elif searchTagNumber == True and searchKunyeNo == False: #Sadece Etiket No aynıysa
            msgboxtext = "'" + chosen_certificate_number + "'" + " sertifikası numarası kullanılıyor.\nFarklı bir numara deneyin!!!"
            QMessageBox.warning(self, 'UYARI', msgboxtext, QMessageBox.Close)

        elif searchTagNumber == True and searchDepartment == True and searchDeviceName == True and searchKunyeNo == True: #Herşey aynıysa
            msgboxtext = chosen_department + "-" + chosen_device + "-" + chosen_certificate_number + " kaydı bulunmaktadır.\nÜstüne yazmak ister misiniz?"
            btnresponse = QMessageBox.question(self, "ONAY", msgboxtext, QMessageBox.Yes | QMessageBox.No )
            if btnresponse == QMessageBox.Yes:
                devamke()

        else:
            devamke()

    def add_record_to_excel_Serial(self):

        if self.ui.checkBoxCommentSerial.isChecked() == True and self.ui.CommentBoxSerial.toPlainText() == "":
            optionalCommand = ""
        elif self.ui.checkBoxCommentSerial.isChecked() == False:
            optionalCommand = ""
        else:
            optionalCommand = "-" + self.ui.CommentBoxSerial.toPlainText()

        date_text = self.ui.dateEdit.text()
        month = date_text[3:5]
        year = date_text[8:10]

        time = QDateTime.currentDateTime()
        chosen_hospital = self.ui.cb_hospital.currentText()
        chosen_birlik = self.ui.line_birlik.text()
        chosen_device = self.ui.cb_device.currentText()
        chosen_department = self.ui.cb_department.currentText()
        chosen_brand = self.ui.cb_brand.currentText()
        chosen_model = self.ui.cb_model.currentText()
        chosen_serialnum = self.ui.seri_no.text()
        chosen_knyeno = self.ui.kn_no.text()
        chosen_date = self.ui.dateEdit.text()
        chosen_calpeson = self.ui.cb_calibpersonSerial.currentText()
        chose_order_number = month + year + self.ui.cb_hospitalcode.currentText() + self.ui.line_isno.text()

        az1Lux = self.ui.lxValue.text()
        az1Sıcaklık = self.ui.celciusValue.text()
        az1MaxPower = self.ui.maxPower.text()

        az2ChargeTime = self.ui.chargeTime.text()

        az3DelayTime = self.ui.delayTime.text()


        if chosen_hospital == "" or chosen_birlik == "" or chosen_device == "" or chosen_department == "" or chosen_brand == "" or chosen_model == "" or chosen_serialnum == "" or chosen_knyeno == "" or chosen_date == "" or chosen_calpeson == "":
           self.statusBar().showMessage("Bilgiler eksik, Gerekli bilgileri giriniz...", 2000) 
        else:
            if self.ui.tag_code_endSerial.text() == "" and self.ui.tag_code_to_endSerial.text() == "":
                self.statusBar().showMessage("Sertfika kayıt aralığı numaralarını giriniz...", 2000)
            elif self.ui.tag_code_endSerial.text() == "":
                self.statusBar().showMessage("Sertfika başlangıç numarasını giriniz...", 2000) 
            elif self.ui.tag_code_to_endSerial.text() == "":
                self.statusBar().showMessage("Sertfika sonlandırma numarasını giriniz...", 2000) 
            else:
                startTagNumber = int(self.ui.tag_code_endSerial.text())
                stopTagNumber = int(self.ui.tag_code_to_endSerial.text())

                with open('pathFolder.txt', 'r', encoding="utf-8") as file:
                    pathSCW = file.read().splitlines()

                for i in range(startTagNumber, stopTagNumber + 1 ):
                    if len(str(i)) == 1:
                        self.ui.tag_code_frontSerial.setText(month + year + self.ui.cb_hospitalcode.currentText() + "000")
                    elif len(str(i)) == 2:
                        self.ui.tag_code_frontSerial.setText(month + year + self.ui.cb_hospitalcode.currentText() + "00")
                    elif len(str(i)) == 3:
                        self.ui.tag_code_frontSerial.setText(month + year + self.ui.cb_hospitalcode.currentText() + "0")
                    elif len(str(i)) == 4:
                        self.ui.tag_code_frontSerial.setText(month + year + self.ui.cb_hospitalcode.currentText() + "")

                    serialTagFrontCode = self.ui.tag_code_frontSerial.text()

                    try:
                        standart_certificate = load_workbook(pathSCW[0]+ "\\" + chosen_device + ".xlsx")
                        worksheet = standart_certificate.active
                        worksheet = standart_certificate["Veri"]
                        worksheet ['O1'] = chosen_hospital
                        worksheet ['O2'] = chosen_birlik
                        worksheet ['O10'] = chosen_department
                        worksheet ['O5'] = chosen_brand
                        worksheet ['O6'] = chosen_model
                        worksheet ['O7'] = chosen_serialnum
                        worksheet ['O11'] = chosen_knyeno
                        worksheet ['O5'] = chosen_brand
                        worksheet ['O8'] = chosen_date
                        worksheet ['O15'] = chosen_calpeson
                        worksheet ['O12'] = serialTagFrontCode + str(i)

                        try:
                            az1List = [az1Lux, az1Sıcaklık, az1MaxPower]
                            az1Value = list(filter(lambda x: x != "", az1List))[0]
                        except IndexError:
                            az1Value = ""
                        
                        try:
                            az2list = [az2ChargeTime]
                            az2Value = list(filter(lambda x: x != "", az2list))[0]
                        except IndexError:
                            az2Value = ""

                        try:
                            az3list = [az3DelayTime]
                            az3Value = list(filter(lambda x: x != "", az3list))[0]
                        except IndexError:
                            az3Value = ""
                        
                        try:
                            worksheet ['AZ1'] = locale.atof(az1Value)
                        except ValueError:
                            pass
                        
                        try:
                            worksheet ['AZ2'] = locale.atof(az2Value)
                        except ValueError:
                            pass

                        try:
                            worksheet ['AZ3'] = locale.atof(az3Value)
                        except ValueError:
                            pass

                        standart_certificate.save(pathSCW[1]+ "\\" + chosen_department + "-" + chosen_device + "-" + serialTagFrontCode + str(i) +  optionalCommand +  ".xlsx")
                        standart_certificate.close()
                        self.ui.statusbar.showMessage(serialTagFrontCode + str(i) + ": Sertifika numaralı kayıtlar eklendi.",500)

                        with open (thisSetupPath + "\\history\\" + chose_order_number +".csv", "a", newline='', encoding='utf-8') as file:
                            csv_writer = writer(file)                  
                            csv_writer.writerow([time.toString(), chosen_department, chosen_device, serialTagFrontCode + str(i), chosen_serialnum, chosen_knyeno])
                            self.ui.listWidget.scrollToBottom()

                        self.ui.add_recordSerial.setEnabled(False)
                        QtTest.QTest.qWait(1000)
                        self.ui.listWidget.clear()

                        with open(thisSetupPath + "\\history\\" + chose_order_number + ".csv", encoding='utf-8') as file:
                            csv_reader = csv.DictReader(file)

                            for row  in csv_reader:
                                tagCode = f'{row["Etiket No"]}'
                                
                                self.ui.listWidget.addItems([tagCode])
                                self.ui.listWidget.scrollToBottom()

                        self.ui.add_recordSerial.setEnabled(True)
                        self.ui.CommentBoxSerial.clear()
                        self.ui.celciusValue.clear()
                        self.ui.lxValue.clear()
                        self.ui.maxPower.clear()
                        self.ui.chargeTime.clear()
                        self.ui.delayTime.clear()
                        
       
                    except FileNotFoundError:
                        msgboxtext = "Standart Sertifika bulunamıyor.\nDosya yolunu kontrol ediniz veya ilgili cihaza ait Standart Sertifika ekleyiniz."
                        QMessageBox.warning(self, 'UYARI', msgboxtext, QMessageBox.Close)
                        break
                        
    def showDetailofData(self):

        date_text = self.ui.dateEdit.text()
        month = date_text[3:5]
        year = date_text[8:10]
        chose_order_number = month + year + self.ui.cb_hospitalcode.currentText() + self.ui.line_isno.text()
        currentItem=str(self.ui.listWidget.currentItem().text())

        with open('pathFolder.txt', 'r', encoding="utf-8") as file:
            pathSCW = file.read().splitlines()

        with open(thisSetupPath + "\\history\\" + chose_order_number +".csv", encoding='utf-8') as file:
            csv_reader = csv.reader(file)
            next(csv_reader)
            for row in csv_reader:
                if currentItem == row[3]:
                    msgboxtext = "Bulunduğu Yer: " + row[1] + "\n"+ "Cihaz Adı: " + row[2] + "\n" + "Seri No: " + row[4] + "\n" + "Künye No: " + row[5]
                    QMessageBox.information(self, currentItem, msgboxtext, QMessageBox.Close)

    def openAddNewCus(self):
        self.addnewcus.show()
    
    def openEditCus(self):
        self.editcurrentcustomer.show()
    
    def openEditStandartCertificateWay(self):
        self.editStandartCertificateWay.show()
    
    def openEditRecordFileWay(self):
        self.editRecordFileWay.show()

    def slideaddRecordFrame(self):
        if self.ui.deviceStackedWidget.currentIndex() == 0 and self.ui.addRecordFrame.y() == 250:
            pass
        elif self.ui.deviceStackedWidget.currentIndex() == 0 and self.ui.addRecordFrame.y() == 310:
            self.positonanimation = QPropertyAnimation(self.ui.addRecordFrame, b"geometry")
            self.positonanimation.setDuration(500)
            self.positonanimation.setStartValue(QRect(10,310,570,125))
            self.positonanimation.setEndValue(QRect(10,250,570,125))
            self.positonanimation.setEasingCurve(QtCore.QEasingCurve.OutQuart)
            self.positonanimation.start()
        elif self.ui.deviceStackedWidget.currentIndex() != 0 and self.ui.addRecordFrame.y() == 250:
            self.positonanimation = QPropertyAnimation(self.ui.addRecordFrame, b"geometry")
            self.positonanimation.setDuration(500)
            self.positonanimation.setStartValue(QRect(10,250,570,125))
            self.positonanimation.setEndValue(QRect(10,310,570,125))
            self.positonanimation.setEasingCurve(QtCore.QEasingCurve.OutQuart)
            self.positonanimation.start()
    
    def setWidthEnterValueFrame(self):
        if self.ui.deviceStackedWidget.currentIndex() == 0 and self.ui.enterValueFrame.height() == 0:
            pass
        elif self.ui.deviceStackedWidget.currentIndex() == 0 and self.ui.enterValueFrame.height() == 60:
            self.heightanimation = QPropertyAnimation(self.ui.enterValueFrame, b"geometry")
            self.heightanimation.setDuration(500)
            self.heightanimation.setStartValue(QRect(10,250,575,60))
            self.heightanimation.setEndValue(QRect(10,250,575,0))
            self.heightanimation.setEasingCurve(QtCore.QEasingCurve.OutQuart)
            self.heightanimation.start()
        elif self.ui.deviceStackedWidget.currentIndex() != 0 and self.ui.enterValueFrame.height() == 0:
            self.heightanimation = QPropertyAnimation(self.ui.enterValueFrame, b"geometry")
            self.heightanimation.setDuration(500)
            self.heightanimation.setStartValue(QRect(10,250,575,0))
            self.heightanimation.setEndValue(QRect(10,250,575,60))
            self.heightanimation.setEasingCurve(QtCore.QEasingCurve.OutQuart)
            self.heightanimation.start()
            


 

def app():
    app = QtWidgets.QApplication(sys.argv)
    app.setStyle("Fusion")
    win = MP_Ui()
    win.show()
    sys.exit(app.exec_())


app()
